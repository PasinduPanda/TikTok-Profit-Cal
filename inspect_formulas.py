import openpyxl
import sys
import io

# Force utf-8 encoding for stdout
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

file_path = "Price & Profit Calculation Template.xlsx"

try:
    wb = openpyxl.load_workbook(file_path, data_only=False)
    
    sheets_to_check = ['Tiktok US', 'Tiktok US ( Profit )', 'Tiktok UK']
    
    for sheet_name in sheets_to_check:
        if sheet_name in wb.sheetnames:
            print(f"\n--- Sheet: {sheet_name} ---")
            ws = wb[sheet_name]
            
            # Print headers
            headers = []
            for cell in ws[1]:
                headers.append(cell.value)
            print(f"Headers: {headers}")
            
            # Check row 2 and 3
            for i in range(2, 4):
                print(f"Row {i} Data/Formulas:")
                row_values = []
                for j, cell in enumerate(ws[i]):
                    val = cell.value
                    # If it's a formula, it starts with =
                    if isinstance(val, str) and val.startswith('='):
                         pass # Keep as is
                    elif isinstance(val, float):
                        val = f"{val:.4f}"
                    
                    # Also print column letter for reference
                    col_letter = openpyxl.utils.get_column_letter(j+1)
                    row_values.append(f"{col_letter}: {val}")
                print(row_values)
                
        else:
            print(f"Sheet '{sheet_name}' not found.")

except Exception as e:
    print(f"Error: {e}")
